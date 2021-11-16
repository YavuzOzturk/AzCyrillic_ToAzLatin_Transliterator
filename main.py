import multiprocessing
import time
from multiprocessing import Process, Queue, Pool
import openpyxl
from pathlib import Path
from os import walk

def write_csv(path, csv_row):
    full_path = 'output_folder' + path
    with open(full_path, 'a', newline='', encoding="utf-8") as f:
        f.write(csv_row)

def read_xlsx(fileName):
    path = 'input_folder'
    xlsx_file = Path(path+fileName)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    # wb = Workbook()
    # ws = wb.active
    # ws.title = "Decoded_Xlsx"
    height = 1
    #wb.save(filename="Decoded.xlsx")
    print(fileName)
    for row in sheet.iter_rows(max_row = sheet.max_row) :
        temp = ""
        row_curr = ""
        cell = 0
        for cell in range(sheet.max_column):
            temp = str(row[cell].value)
            # Аа
            temp = temp.replace("А", "A")
            temp = temp.replace("а", "a")
            # Бб
            temp = temp.replace("Б", "B")
            temp = temp.replace("б", "b")
            # Вв
            temp = temp.replace("В", "V")
            temp = temp.replace("в", "v")
            # Гг
            temp = temp.replace("Г", "Q")
            temp = temp.replace("г", "q")
            # Дд
            temp = temp.replace("Д", "D")
            temp = temp.replace("д", "d")
            # Ее
            temp = temp.replace("Е", "E")
            temp = temp.replace("е", "e")
            # Ёё
            # Жж
            temp = temp.replace("Ж", "C")
            temp = temp.replace("ж", "c")
            # Зз
            temp = temp.replace("З", "Z")
            temp = temp.replace("з", "z")
            # Ии
            temp = temp.replace("И", "İ")
            temp = temp.replace("и", "i")
            # Йй
            temp = temp.replace("Й", "Y")
            temp = temp.replace("й", "y")
            # Кк
            temp = temp.replace("К", "K")
            temp = temp.replace("к", "k")
            # Лл
            temp = temp.replace("Л", "L")
            temp = temp.replace("л", "l")
            # Мм
            temp = temp.replace("М", "M")
            temp = temp.replace("м", "m")
            # Нн
            temp = temp.replace("Н", "N")
            temp = temp.replace("н", "n")
            # Оо
            temp = temp.replace("О", "O")
            temp = temp.replace("о", "o")
            # Пп
            temp = temp.replace("П", "P")
            temp = temp.replace("п", "p")
            # Рр
            temp = temp.replace("Р", "R")
            temp = temp.replace("р", "r")
            # Сс
            temp = temp.replace("С", "S")
            temp = temp.replace("с", "s")
            # Тт
            temp = temp.replace("Т", "T")
            temp = temp.replace("т", "t")
            # Уу
            temp = temp.replace("У", "U")
            temp = temp.replace("у", "u")
            # Фф
            temp = temp.replace("Ф", "F")
            temp = temp.replace("ф", "f")
            # Хх
            temp = temp.replace("Х", "X")
            temp = temp.replace("х", "x")
            # Цц
            temp = temp.replace("Ц", "Ü")
            temp = temp.replace("ц", "ü")
            # Чч
            temp = temp.replace("Ч", "Ç")
            temp = temp.replace("ч", "ç")
            # Шш
            temp = temp.replace("Ш", "Ş")
            temp = temp.replace("ш", "ş")
            # Щщ
            temp = temp.replace("Щ", "H")
            temp = temp.replace("щ", "h")
            # Ъъ
            temp = temp.replace("Ъ", "J")
            temp = temp.replace("ъ", "j")
            # Ыы
            temp = temp.replace("Ы", "I")
            temp = temp.replace("ы", "ı")
            # Ьь
            temp = temp.replace("Ь", "Ğ")
            temp = temp.replace("ь", "ğ")
            # Ээ
            temp = temp.replace("Э", "G")
            temp = temp.replace("э", "g")
            # Юю
            temp = temp.replace("Ю", "Ö")
            temp = temp.replace("ю", "ö")
            # Яя
            temp = temp.replace("Я", "Ə")
            temp = temp.replace("я", "ə")
            # ws[chr(ord('A')+cell)+str(height)] = temp
            # print(ws[chr(ord('A')+cell)+str(height)].value)
            row_curr += temp + " , "
            cell += 1
        #print(row_curr[:-3])
        write_csv(fileName.replace(".xlsx", ".csv"), (row_curr[:-2] + "\n"))
        height += 1
    # wb.save(filename=(path+"/decoded/"+fileName))

def print_test(path):
    time.sleep(3)
    print(path)

def create_queue(path, q):
    filenames_s = next(walk(path), (None, None, []))[2]
    for file in filenames_s:
        q.put(file, True, None)

if __name__ == '__main__':
    #multiprocess addition
    q = Queue()
    procs = []
    create_queue('input_folder', q)
    pool = multiprocessing.Pool(processes=(multiprocessing.cpu_count()-1))

    while not (q.empty()):
        # proc_f_1 = pool.map_async(search_student, q)
        # proc_f_1 = Process(target=read_xlsx, args=(path_src, q.get()))
        res = pool.apply_async(read_xlsx, args=(q.get(),))

        # complete the processes
    pool.close()
    pool.join()

    # filenames = next(walk('input_folder'), (None, None, []))[2]  # [] if no file
    # print(filenames)
    # for fileName in filenames:
    #     print(fileName.replace(".xlsx", ".csv"))
    #     read_xlsx('excel_files/', fileName)

